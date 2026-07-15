from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest

from reports.research_excel import build_combined_workbook

# Config real do projeto -- reaproveitada como está, não recriada em
# miniatura, para o teste exercitar a MESMA ligação feature->coluna que o
# pipeline de produção usa.
MODEL_PATH = Path("config/model.yaml")
DEAL_BREAKERS_PATH = Path("config/deal_breakers.json")


def _observation(symbol: str, name: str, **overrides) -> dict:
    base = {
        "symbol": symbol,
        "name": name,
        "sector": "Technology",
        "quote_type": "EQUITY",
        "currency": "USD",
        "country": "United States",
        "price": 100.0,
        "market_cap": 5_000_000_000.0,
        "volume": 1_000_000.0,
        "roic": 0.20,
        "roe": 0.25,
        "gross_margin": 0.55,
        "pe": 18.0,
        "current_ratio": 1.5,
        "rsi_14": 55.0,
    }
    base.update(overrides)
    return base


def _write_collection(path: Path, observations: dict[str, dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps({"observations": observations}),
        encoding="utf-8",
    )


def _ranking(companies: list[dict]) -> dict:
    return {
        "generated_at": "2026-07-14T00:00:00",
        "summary": {
            "total_count": len(companies),
            "universe_eligible_count": len(companies),
            "candidate_count": sum(1 for c in companies if c.get("safeguard_passed")),
            "blocked_by_reason": {},
        },
        "companies": companies,
    }


def _company(symbol: str, rank: int, candidate: bool = True) -> dict:
    return {
        "symbol": symbol,
        "sector": "Technology",
        "universe_eligible": True,
        "safeguard_passed": candidate,
        "safeguard_reasons": [] if candidate else ["CONFIDENCE_BELOW_MINIMUM"],
        "market_rank": rank,
        "candidate_rank": rank if candidate else None,
        "investment_score": 70.0,
        "opportunity_score": 60.0,
        "conviction_score": 65.0,
        "confidence_score": 90.0,
        "already_held": False,
    }


def _portfolio() -> dict:
    return {
        "summary": {"position_count": 1, "invested_weight": 0.05},
        "positions": [
            {
                "candidate_rank": 1, "symbol": "AAA", "name": "Alpha",
                "sector": "Technology", "industry": "Software",
                "target_weight": 0.05, "investment_score": 70.0,
                "reference_price": 100.0,
            }
        ],
    }


@pytest.fixture
def fake_workspace(tmp_path: Path) -> Path:
    output_dir = tmp_path / "output"
    data_dir = output_dir / "dados"
    data_dir.mkdir(parents=True)
    observations = {
        "AAA": _observation("AAA", "Alpha", roic=0.30),
        "BBB": _observation("BBB", "Beta", roic=0.05),
    }
    _write_collection(tmp_path / "collection.json", observations)

    ranking = _ranking([_company("AAA", 1, True), _company("BBB", 2, False)])
    (data_dir / "research_ranking_report.json").write_text(
        json.dumps(ranking), encoding="utf-8"
    )
    (data_dir / "model_portfolio_report.json").write_text(
        json.dumps(_portfolio()), encoding="utf-8"
    )
    return tmp_path


def test_combined_workbook_exposes_factor_and_feature_detail(
    fake_workspace: Path,
) -> None:
    output_dir = fake_workspace / "output"
    # ROOT / caminho_absoluto devolve o caminho absoluto inalterado (regra
    # padrão do pathlib) -- um caminho absoluto aqui funciona sem precisar
    # que a coleção esteja de fato dentro do ROOT do projeto.
    workbook_path = build_combined_workbook(
        (("", "S&P 500", str(fake_workspace / "collection.json")),),
        output_dir,
        output_dir / "combined.xlsx",
        model_path=MODEL_PATH,
        deal_breakers_path=DEAL_BREAKERS_PATH,
    )

    detail = pd.read_excel(workbook_path, sheet_name="Todos os Itens")
    assert len(detail) == 2
    assert "Business: ROIC (valor)" in detail.columns
    assert "Business: ROIC (percentil)" in detail.columns
    assert "Investment Score" in detail.columns

    aaa = detail[detail["Símbolo"] == "AAA"].iloc[0]
    bbb = detail[detail["Símbolo"] == "BBB"].iloc[0]
    # AAA tem ROIC bruto maior -> percentil maior no mesmo lote.
    assert aaa["Business: ROIC (valor)"] > bbb["Business: ROIC (valor)"]
    assert aaa["Business: ROIC (percentil)"] > bbb["Business: ROIC (percentil)"]
    assert aaa["Status"] == "Candidato #1"
    assert bbb["Status"] == "Bloqueado: CONFIDENCE_BELOW_MINIMUM"


def test_missing_screener_is_omitted_not_an_error(fake_workspace: Path) -> None:
    output_dir = fake_workspace / "output"
    workbook_path = build_combined_workbook(
        (
            ("", "S&P 500", str(fake_workspace / "collection.json")),
            ("market", "Mercado Amplo", str(fake_workspace / "does_not_exist.json")),
        ),
        output_dir,
        output_dir / "combined.xlsx",
        model_path=MODEL_PATH,
        deal_breakers_path=DEAL_BREAKERS_PATH,
    )
    summary = pd.read_excel(workbook_path, sheet_name="Resumo")
    assert list(summary["Screener"]) == ["S&P 500"]


def test_no_screener_collected_raises(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError):
        build_combined_workbook(
            (("", "S&P 500", "does-not-matter.json"),),
            tmp_path,
            tmp_path / "combined.xlsx",
            model_path=MODEL_PATH,
            deal_breakers_path=DEAL_BREAKERS_PATH,
        )


def test_column_groups_are_applied(fake_workspace: Path) -> None:
    import openpyxl

    output_dir = fake_workspace / "output"
    workbook_path = build_combined_workbook(
        (("", "S&P 500", str(fake_workspace / "collection.json")),),
        output_dir,
        output_dir / "combined.xlsx",
        model_path=MODEL_PATH,
        deal_breakers_path=DEAL_BREAKERS_PATH,
    )
    wb = openpyxl.load_workbook(workbook_path)
    worksheet = wb["Todos os Itens"]
    headers = {
        worksheet.cell(row=1, column=col).value: col
        for col in range(1, worksheet.max_column + 1)
    }

    symbol_column = worksheet.column_dimensions[
        worksheet.cell(row=1, column=headers["Símbolo"]).column_letter
    ]
    assert symbol_column.outline_level == 0
    assert not symbol_column.hidden

    factor_column = worksheet.column_dimensions[
        worksheet.cell(row=1, column=headers["Business Score"]).column_letter
    ]
    assert factor_column.outline_level == 1
    assert factor_column.hidden

    feature_column = worksheet.column_dimensions[
        worksheet.cell(row=1, column=headers["Business: ROIC (valor)"]).column_letter
    ]
    assert feature_column.outline_level == 2
    assert feature_column.hidden
