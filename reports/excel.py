from __future__ import annotations

from datetime import datetime
from pathlib import Path
import shutil

import pandas as pd
from openpyxl.styles import Font, PatternFill

from reports.diagnostics import build_diagnostics
from reports.explainability import build_explainability
from reports.history_report import (
    build_historical_trends,
    build_history_summary,
)
from reports.report_engine import build_company_reports
from reports.report_models import CompanyReport
from portfolio.report import PortfolioReport



def _join_items(items: tuple[str, ...]) -> str:
    return "; ".join(items)


def _company_reports_dataframe(
    reports: list[CompanyReport],
) -> pd.DataFrame:
    """
    Converte objetos CompanyReport em uma tabela de apresentação.

    Esta função isola o Excel do DataFrame bruto do pipeline para
    a aba Decision Analysis.
    """

    rows: list[dict[str, object]] = []

    for report in reports:
        rows.append(
            {
                "symbol": report.symbol,
                "name": report.company_name,
                "Decision": report.decision,
                "Decision Rating": report.decision_rating,
                "Suggested Action": report.suggested_action,
                "Decision Confidence": report.decision_confidence,
                "Decision Drivers": _join_items(
                    report.decision_drivers
                ),
                "Investment Thesis": report.investment_thesis,
                "Thesis Strengths": _join_items(report.strengths),
                "Thesis Risks": _join_items(report.risks),
                "Thesis Catalysts": _join_items(report.catalysts),
                "Opportunity Score": report.opportunity_score,
                "Conviction Score": report.conviction_score,
                "Investment Score": report.investment_score,
                "Business Score": report.business_score,
                "Valuation Score": report.valuation_score,
                "Financial Score": report.financial_score,
                "Timing Score": report.timing_score,
                "Confidence Score": report.confidence_score,
                "Risk Penalty": report.risk_penalty,
                "Deal Breakers": _join_items(
                    report.deal_breakers
                ),
            }
        )

    result = pd.DataFrame(rows)

    if result.empty:
        return result

    sort_columns = [
        column
        for column in [
            "Opportunity Score",
            "Conviction Score",
            "Investment Score",
        ]
        if column in result.columns
    ]

    if sort_columns:
        result = result.sort_values(
            sort_columns,
            ascending=[False] * len(sort_columns),
            na_position="last",
        )

    return result.reset_index(drop=True)


def _portfolio_summary_dataframe(report: PortfolioReport) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Metric": str(key).replace("_", " ").title(),
                "Value": value,
            }
            for key, value in report.summary.items()
        ]
    )


def _portfolio_allocation_dataframe(report: PortfolioReport) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    labels = {
        "by_symbol": "Symbol",
        "by_sector": "Sector",
        "by_country": "Country",
        "by_currency": "Currency",
    }
    for key, dimension in labels.items():
        values = report.allocation.get(key, {})
        if isinstance(values, dict):
            for name, weight in values.items():
                rows.append(
                    {
                        "Dimension": dimension,
                        "Name": name,
                        "Weight": weight,
                    }
                )
    rows.append(
        {
            "Dimension": "Cash",
            "Name": "Cash",
            "Weight": report.allocation.get("cash_weight", 0.0),
        }
    )
    return pd.DataFrame(rows)


def _portfolio_concentration_dataframe(report: PortfolioReport) -> pd.DataFrame:
    rows = [
        {"Metric": "Concentration Score", "Value": report.concentration.get("concentration_score")},
        {"Metric": "Diversification Score", "Value": report.concentration.get("diversification_score")},
        {"Metric": "Largest Position Weight", "Value": report.concentration.get("largest_position_weight")},
        {"Metric": "Top 5 Weight", "Value": report.concentration.get("top_5_weight")},
    ]
    dimensions = {
        "sector_concentration": "Sector",
        "country_concentration": "Country",
        "currency_concentration": "Currency",
    }
    for key, label in dimensions.items():
        values = report.concentration.get(key, {})
        if isinstance(values, dict):
            for name, weight in values.items():
                rows.append(
                    {
                        "Metric": f"{label}: {name}",
                        "Value": weight,
                    }
                )
    return pd.DataFrame(rows)


def _portfolio_quality_dataframe(report: PortfolioReport) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Metric": str(key).replace("_", " ").title(),
                "Value": value,
            }
            for key, value in report.quality.items()
            if key not in {"warnings", "missing_report_symbols"}
        ]
    )


def _portfolio_rebalance_dataframe(report: PortfolioReport) -> pd.DataFrame:
    actions = report.rebalance.get("actions", [])
    if not isinstance(actions, list):
        return pd.DataFrame()
    return pd.DataFrame(actions)


def _format_portfolio_sheet(
    writer: pd.ExcelWriter,
    sheet_name: str,
) -> None:
    worksheet = writer.sheets.get(sheet_name)

    if worksheet is None:
        return

    for cell in worksheet[1]:
        cell.font = Font(
            name=cell.font.name or "Calibri",
            size=cell.font.size or 11,
            bold=True,
            color="FFFFFF",
        )
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

    percentage_metrics = {
        "Cash Weight",
        "Covered Weight",
        "Largest Position Weight",
        "Top 5 Weight",
        "Estimated Turnover",
    }
    money_metrics = {
        "Total Market Value",
        "Cash",
        "Total Value",
        "Required Cash",
        "Released Cash",
        "Net Cash Requirement",
    }

    if sheet_name in {"Portfolio Summary", "Portfolio Quality"}:
        for row in range(2, worksheet.max_row + 1):
            metric = str(worksheet.cell(row=row, column=1).value or "")
            value_cell = worksheet.cell(row=row, column=2)
            if metric in percentage_metrics:
                value_cell.number_format = "0.0%"
            elif metric in money_metrics:
                value_cell.number_format = '#,##0.00;[Red](#,##0.00);-'
            elif "Score" in metric:
                value_cell.number_format = "0.0"

    elif sheet_name == "Portfolio Allocation":
        for cell in worksheet["C"][1:]:
            cell.number_format = "0.0%"

    elif sheet_name == "Portfolio Concentration":
        for row in range(2, worksheet.max_row + 1):
            metric = str(worksheet.cell(row=row, column=1).value or "")
            value_cell = worksheet.cell(row=row, column=2)
            value_cell.number_format = (
                "0.0"
                if metric in {
                    "Concentration Score",
                    "Diversification Score",
                }
                else "0.0%"
            )

    elif sheet_name == "Portfolio Rebalance":
        headers = {
            cell.value: cell.column
            for cell in worksheet[1]
        }
        for header in {"current_weight", "target_weight"}:
            column = headers.get(header)
            if column is not None:
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=column).number_format = (
                        "0.0%"
                    )
        for header in {"target_value", "trade_value"}:
            column = headers.get(header)
            if column is not None:
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=column).number_format = (
                        '#,##0.00;[Red](#,##0.00);-'
                    )

    worksheet.sheet_view.showGridLines = False

def _format_sheet(
    writer: pd.ExcelWriter,
    sheet_name: str,
) -> None:
    """
    Aplica formatação básica e reutilizável às abas do Excel.
    """

    worksheet = writer.sheets.get(sheet_name)

    if worksheet is None:
        return

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = cell.value

            if value is not None:
                max_length = max(
                    max_length,
                    len(str(value)),
                )

        worksheet.column_dimensions[column_letter].width = min(
            max_length + 2,
            40,
        )


def _build_historical_reports(
    database_path: Path,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Gera os relatórios históricos sem interromper a criação
    do Excel caso o banco ainda não exista ou esteja vazio.
    """

    if not database_path.exists():
        return pd.DataFrame(), pd.DataFrame()

    try:
        trends = build_historical_trends(
            database_path=database_path,
            period_days=30,
        )

        summary = build_history_summary(
            database_path=database_path,
        )

        return trends, summary

    except Exception as exc:
        print(
            "[AVISO] Não foi possível gerar os relatórios históricos: "
            f"{exc}"
        )

        return pd.DataFrame(), pd.DataFrame()


def write_latest_and_history(
    df: pd.DataFrame,
    output_dir: Path,
    portfolio_report: PortfolioReport | None = None,
) -> tuple[Path, Path | None]:
    """
    Gera o arquivo Excel histórico da execução e atualiza latest.xlsx.

    Abas geradas:

    - Ranking
    - Summary
    - Opportunity Analysis
    - Decision Analysis
    - Explainability
    - Diagnostics
    - Historical Trends
    - History Summary
    - Portfolio Summary (quando disponível)
    - Portfolio Allocation (quando disponível)
    - Portfolio Concentration (quando disponível)
    - Portfolio Quality (quando disponível)
    - Portfolio Rebalance (quando disponível)
    - Portfolio Warnings (quando disponível)
    """

    output_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    history_dir = output_dir / "history"

    history_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    root_dir = output_dir.parent
    database_path = root_dir / "data" / "atlas_history.db"

    timestamp = datetime.now().strftime(
        "%Y-%m-%d_%H-%M-%S"
    )

    history_file = (
        history_dir
        / f"atlas_snapshot_{timestamp}.xlsx"
    )

    latest_file = output_dir / "latest.xlsx"

    summary_columns = [
        column
        for column in [
            "symbol",
            "name",
            "Investment Score",
            "Opportunity Score",
            "Opportunity Rating",
            "Conviction Score",
            "Conviction Rating",
            "Decision",
            "Decision Rating",
            "Suggested Action",
            "Decision Confidence",
            "Business Score",
            "Valuation Score",
            "Financial Score",
            "Timing Score",
            "Confidence Score",
            "Risk Penalty",
            "Recommendation",
        ]
        if column in df.columns
    ]

    opportunity_columns = [
        column
        for column in [
            "symbol",
            "name",
            "Opportunity Base",
            "Opportunity Bonus",
            "Opportunity Penalty",
            "Opportunity Score",
            "Opportunity Rating",
            "Opportunity Drivers",
            "Investment Score",
            "Business Score",
            "Valuation Score",
            "Financial Score",
            "Timing Score",
            "Confidence Score",
            "Risk Penalty",
            "Deal Breakers",
        ]
        if column in df.columns
    ]

    decision_columns = [
        column
        for column in [
            "symbol",
            "name",
            "Decision",
            "Decision Rating",
            "Suggested Action",
            "Decision Confidence",
            "Decision Drivers",
            "Investment Thesis",
            "Thesis Strengths",
            "Thesis Risks",
            "Thesis Catalysts",
            "Opportunity Score",
            "Conviction Score",
            "Investment Score",
            "Business Score",
            "Valuation Score",
            "Financial Score",
            "Timing Score",
            "Confidence Score",
            "Risk Penalty",
            "Deal Breakers",
        ]
        if column in df.columns
    ]

    company_reports = build_company_reports(df)
    decision_analysis_df = _company_reports_dataframe(
        company_reports
    )

    explainability_df = build_explainability(df)
    diagnostics_df = build_diagnostics(df)

    historical_trends_df, history_summary_df = (
        _build_historical_reports(database_path)
    )

    with pd.ExcelWriter(
        history_file,
        engine="openpyxl",
    ) as writer:

        # --------------------------------------------------------------
        # Ranking
        # --------------------------------------------------------------
        df.to_excel(
            writer,
            sheet_name="Ranking",
            index=False,
        )

        # --------------------------------------------------------------
        # Summary
        # --------------------------------------------------------------
        if summary_columns:
            df[summary_columns].to_excel(
                writer,
                sheet_name="Summary",
                index=False,
            )

        # --------------------------------------------------------------
        # Opportunity Analysis
        # --------------------------------------------------------------
        if opportunity_columns:
            opportunity_df = df[
                opportunity_columns
            ].copy()

            if "Opportunity Score" in opportunity_df.columns:
                opportunity_df = opportunity_df.sort_values(
                    "Opportunity Score",
                    ascending=False,
                    na_position="last",
                )

            opportunity_df.to_excel(
                writer,
                sheet_name="Opportunity Analysis",
                index=False,
            )


        # --------------------------------------------------------------
        # Decision Analysis
        # --------------------------------------------------------------
        if not decision_analysis_df.empty:
            decision_analysis_df.to_excel(
                writer,
                sheet_name="Decision Analysis",
                index=False,
            )

        # --------------------------------------------------------------
        # Explainability
        # --------------------------------------------------------------
        explainability_df.to_excel(
            writer,
            sheet_name="Explainability",
            index=False,
        )

        # --------------------------------------------------------------
        # Diagnostics
        # --------------------------------------------------------------
        if not diagnostics_df.empty:
            diagnostics_df.to_excel(
                writer,
                sheet_name="Diagnostics",
                index=False,
            )

        # --------------------------------------------------------------
        # Historical Trends
        # --------------------------------------------------------------
        if not historical_trends_df.empty:
            historical_trends_df.to_excel(
                writer,
                sheet_name="Historical Trends",
                index=False,
            )

        # --------------------------------------------------------------
        # History Summary
        # --------------------------------------------------------------
        if not history_summary_df.empty:
            history_summary_df.to_excel(
                writer,
                sheet_name="History Summary",
                index=False,
            )

        # --------------------------------------------------------------
        # Portfolio Intelligence
        # --------------------------------------------------------------
        if portfolio_report is not None:
            _portfolio_summary_dataframe(portfolio_report).to_excel(
                writer, sheet_name="Portfolio Summary", index=False
            )
            _portfolio_allocation_dataframe(portfolio_report).to_excel(
                writer, sheet_name="Portfolio Allocation", index=False
            )
            _portfolio_concentration_dataframe(portfolio_report).to_excel(
                writer, sheet_name="Portfolio Concentration", index=False
            )
            _portfolio_quality_dataframe(portfolio_report).to_excel(
                writer, sheet_name="Portfolio Quality", index=False
            )
            rebalance_df = _portfolio_rebalance_dataframe(portfolio_report)
            if not rebalance_df.empty:
                rebalance_df.to_excel(
                    writer, sheet_name="Portfolio Rebalance", index=False
                )
            if portfolio_report.warnings:
                pd.DataFrame(
                    {"Warning": list(portfolio_report.warnings)}
                ).to_excel(
                    writer, sheet_name="Portfolio Warnings", index=False
                )

        # --------------------------------------------------------------
        # Formatting
        # --------------------------------------------------------------
        for sheet_name in writer.sheets:
            _format_sheet(writer, sheet_name)
            if sheet_name.startswith("Portfolio "):
                _format_portfolio_sheet(writer, sheet_name)

    copied_file: Path | None = None

    try:
        shutil.copy2(
            history_file,
            latest_file,
        )

        copied_file = latest_file

    except PermissionError:
        print(
            "[AVISO] latest.xlsx está aberto. "
            "O arquivo histórico foi salvo normalmente."
        )

    return history_file, copied_file