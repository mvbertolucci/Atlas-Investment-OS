from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill

from analytics.feature_audit import collect_model_features
from analytics.mapper import normalize_columns
from reports.research_common import company_status
from scoring.investment import score_dataframe

ROOT = Path(__file__).resolve().parent.parent

# (sufixo do arquivo, nome de exibição, coleção bruta correspondente) --
# Mercado Amplo e ADR reaproveitam a MESMA coleção (diferem só na política
# de universo já aplicada no research_ranking_report*.json de cada um);
# rodar score_dataframe nela uma única vez e reaproveitar para os dois.
DEFAULT_SCREENERS: tuple[tuple[str, str, str], ...] = (
    ("", "S&P 500", "data/research_universe_collection.json"),
    ("market", "Mercado Amplo", "data/research_universe_collection_market.json"),
    ("adr", "ADR", "data/research_universe_collection_market.json"),
)

FACTORS = ("business", "valuation", "financial", "timing")

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
_GROUP_FILL = PatternFill(fill_type="solid", fgColor="DCE6F1")
_PERCENT_SUFFIXES = ("Peso investido", "Peso")


def _safe_label(label: str) -> str:
    """Mesma transformação de factors/engine.py::score_factor para nomear
    as colunas de detalhe (ex.: 'Debt / Equity' -> 'Debt___Equity')."""
    return (
        label.replace("/", "_")
        .replace(" ", "_")
        .replace("(", "")
        .replace(")", "")
        .replace("-", "_")
    )


def _factor_feature_bindings(
    features_path: Path, model_path: Path
) -> dict[str, list[tuple[str, str, str]]]:
    """factor -> [(rótulo, coluna_bruta, coluna_de_score), ...] -- a mesma
    ligação feature->coluna que o motor de scoring usa de verdade
    (analytics.feature_audit.collect_model_features), não uma lista
    reinventada."""
    bindings = collect_model_features(features_path, model_path)
    result: dict[str, list[tuple[str, str, str]]] = {factor: [] for factor in FACTORS}
    for binding in bindings:
        if binding.factor not in result:
            continue
        score_column = f"{binding.factor}_{_safe_label(binding.label)}_score"
        result[binding.factor].append((binding.label, binding.column, score_column))
    return result


def _labeled_path(output_dir: Path, base_name: str, label: str) -> Path:
    if not label:
        return output_dir / base_name
    stem, _, suffix = base_name.rpartition(".")
    return output_dir / f"{stem}_{label}.{suffix}"


def _load_json(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def score_collection(
    collection_path: Path,
    model_path: Path,
    deal_breakers_path: Path,
) -> pd.DataFrame:
    """
    Roda o MESMO pipeline determinístico que já produziu o Investment/
    Opportunity/Conviction Score persistidos em research_ranking_report*.json
    (scoring.investment.score_dataframe, chamado igual a
    portfolio.model_portfolio) -- mas desta vez sobre a coleção bruta
    completa, para capturar as colunas intermediárias (fator por fator,
    feature por feature) que o pipeline de ranking descarta depois de
    extrair só o Investment Score final.
    """
    state = _load_json(collection_path)
    if state is None:
        raise FileNotFoundError(f"Coleção não encontrada: {collection_path}")
    features_path = model_path.parent / "features.yaml"
    frame = normalize_columns(pd.DataFrame(list(state["observations"].values())))
    return score_dataframe(frame, model_path, deal_breakers_path)


def _row_and_groups(
    screener_name: str,
    ranking: dict[str, Any],
    scored: pd.DataFrame,
    bindings_by_factor: dict[str, list[tuple[str, str, str]]],
) -> tuple[list[dict[str, Any]], list[tuple[str, int]]]:
    """
    Monta as linhas + o plano de colunas (nome, nível de agrupamento: 0 =
    sempre visível, 1 = agrupado/colapsado, 2 = aninhado dentro do grupo do
    fator). O plano é o mesmo para toda linha -- construído uma vez.
    """
    scored_by_symbol = {
        str(row["symbol"]).strip().upper(): row for _, row in scored.iterrows()
    }
    plan: list[tuple[str, int]] = [
        ("Screener", 0),
        ("Rank", 0),
        ("Símbolo", 0),
        ("Nome", 0),
        ("Setor", 0),
        ("Status", 0),
        ("Já detida", 0),
        ("Investment Score", 0),
        ("Opportunity Score", 0),
        ("Conviction Score", 0),
        ("Confidence Score", 0),
        ("Decision Rating", 0),
        ("Suggested Action", 0),
        ("Risk Penalty", 0),
        ("Deal Breakers", 0),
    ]
    for factor in FACTORS:
        title = factor.title()
        plan.append((f"{title} Score", 1))
        plan.append((f"{title} Confidence", 1))
        for label, _, _ in bindings_by_factor[factor]:
            plan.append((f"{title}: {label} (valor)", 2))
            plan.append((f"{title}: {label} (percentil)", 2))
    plan += [
        ("Opportunity Base", 1),
        ("Opportunity Bonus", 1),
        ("Opportunity Penalty", 1),
        ("Opportunity Rating", 1),
        ("Opportunity Drivers", 1),
        ("Conviction Data Quality", 1),
        ("Conviction Factor Agreement", 1),
        ("Conviction Historical Stability", 1),
        ("Conviction Risk Profile", 1),
        ("Conviction Base", 1),
        ("Conviction Bonus", 1),
        ("Conviction Penalty", 1),
        ("Conviction Rating", 1),
        ("Conviction Drivers", 1),
        ("Decision Confidence", 1),
        ("Decision Priority", 1),
        ("Decision Drivers", 1),
        ("Investment Thesis", 1),
        ("Thesis Strengths", 1),
        ("Thesis Risks", 1),
        ("Thesis Catalysts", 1),
    ]

    rows: list[dict[str, Any]] = []
    for company in ranking.get("companies", []):
        symbol = str(company.get("symbol", "")).strip().upper()
        scored_row = scored_by_symbol.get(symbol)
        if scored_row is None:
            continue

        status_label, _ = company_status(company)
        row: dict[str, Any] = {
            "Screener": screener_name,
            "Rank": company.get("market_rank"),
            "Símbolo": symbol,
            "Nome": scored_row.get("name"),
            "Setor": company.get("sector"),
            "Status": status_label,
            "Já detida": bool(company.get("already_held")),
            "Investment Score": scored_row.get("Investment Score"),
            "Opportunity Score": scored_row.get("Opportunity Score"),
            "Conviction Score": scored_row.get("Conviction Score"),
            "Confidence Score": scored_row.get("Confidence Score"),
            "Decision Rating": scored_row.get("Decision Rating"),
            "Suggested Action": scored_row.get("Suggested Action"),
            "Risk Penalty": scored_row.get("Risk Penalty"),
            "Deal Breakers": scored_row.get("Deal Breakers"),
        }
        for factor in FACTORS:
            title = factor.title()
            row[f"{title} Score"] = scored_row.get(f"{title} Factor")
            row[f"{title} Confidence"] = scored_row.get(f"{title} Confidence")
            for label, raw_column, score_column in bindings_by_factor[factor]:
                row[f"{title}: {label} (valor)"] = scored_row.get(raw_column)
                row[f"{title}: {label} (percentil)"] = scored_row.get(score_column)
        row.update(
            {
                "Opportunity Base": scored_row.get("Opportunity Base"),
                "Opportunity Bonus": scored_row.get("Opportunity Bonus"),
                "Opportunity Penalty": scored_row.get("Opportunity Penalty"),
                "Opportunity Rating": scored_row.get("Opportunity Rating"),
                "Opportunity Drivers": scored_row.get("Opportunity Drivers"),
                "Conviction Data Quality": scored_row.get("Conviction Data Quality"),
                "Conviction Factor Agreement": scored_row.get(
                    "Conviction Factor Agreement"
                ),
                "Conviction Historical Stability": scored_row.get(
                    "Conviction Historical Stability"
                ),
                "Conviction Risk Profile": scored_row.get("Conviction Risk Profile"),
                "Conviction Base": scored_row.get("Conviction Base"),
                "Conviction Bonus": scored_row.get("Conviction Bonus"),
                "Conviction Penalty": scored_row.get("Conviction Penalty"),
                "Conviction Rating": scored_row.get("Conviction Rating"),
                "Conviction Drivers": scored_row.get("Conviction Drivers"),
                "Decision Confidence": scored_row.get("Decision Confidence"),
                "Decision Priority": scored_row.get("Decision Priority"),
                "Decision Drivers": scored_row.get("Decision Drivers"),
                "Investment Thesis": scored_row.get("Investment Thesis"),
                "Thesis Strengths": scored_row.get("Thesis Strengths"),
                "Thesis Risks": scored_row.get("Thesis Risks"),
                "Thesis Catalysts": scored_row.get("Thesis Catalysts"),
            }
        )
        rows.append(row)

    return rows, plan


def _positions_dataframe(portfolio: dict[str, Any], screener_name: str) -> pd.DataFrame:
    rows = [
        {
            "Screener": screener_name,
            "Rank": position.get("candidate_rank"),
            "Símbolo": position.get("symbol"),
            "Nome": position.get("name"),
            "Setor": position.get("sector"),
            "Indústria": position.get("industry"),
            "Peso": position.get("target_weight"),
            "Investment Score": position.get("investment_score"),
            "Preço ref.": position.get("reference_price"),
        }
        for position in portfolio.get("positions", [])
    ]
    return pd.DataFrame(rows)


def _summary_row(
    screener_name: str, ranking: dict[str, Any], portfolio: dict[str, Any] | None
) -> dict[str, Any]:
    summary = ranking.get("summary", {})
    row: dict[str, Any] = {
        "Screener": screener_name,
        "Gerado em": ranking.get("generated_at"),
        "Analisados": summary.get("total_count", 0),
        "Elegíveis": summary.get("universe_eligible_count", 0),
        "Candidatos": summary.get("candidate_count", 0),
    }
    for reason, count in (summary.get("blocked_by_reason") or {}).items():
        row[f"Bloq.: {reason}"] = count
    if portfolio is not None:
        portfolio_summary = portfolio.get("summary", {})
        row["Posições na carteira"] = portfolio_summary.get("position_count", 0)
        row["Peso investido"] = portfolio_summary.get("invested_weight", 0.0)
    return row


def _style_flat_sheet(
    writer: pd.ExcelWriter, sheet_name: str, percent_columns: tuple[str, ...] = ()
) -> None:
    worksheet = writer.sheets.get(sheet_name)
    if worksheet is None or worksheet.max_row < 1:
        return

    for cell in worksheet[1]:
        cell.font = Font(
            name=cell.font.name or "Calibri", size=cell.font.size or 11,
            bold=True, color="FFFFFF",
        )
        cell.fill = _HEADER_FILL

    headers = {cell.value: cell.column for cell in worksheet[1]}
    for header in percent_columns:
        column = headers.get(header)
        if column is None:
            continue
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=column).number_format = "0.0%"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    for column_cells in worksheet.columns:
        max_length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=0,
        )
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(
            max_length + 2, 40
        )


def _style_grouped_sheet(
    writer: pd.ExcelWriter, sheet_name: str, plan: list[tuple[str, int]]
) -> None:
    """
    Estiliza a aba de detalhe: cabeçalho, autofiltro, freeze panes, e o
    agrupamento de colunas em si -- nível 1 (fatores/composições, colapsado)
    e nível 2 (feature a feature dentro de cada fator, colapsado). Clicar no
    "+" no topo da planilha abre o nível correspondente -- é o "abrir os
    cálculos" pedido: o valor final fica visível, o detalhe fica a um clique.
    """
    worksheet = writer.sheets.get(sheet_name)
    if worksheet is None or worksheet.max_row < 1:
        return

    worksheet.sheet_properties.outlinePr.summaryRight = True

    for cell in worksheet[1]:
        cell.font = Font(
            name=cell.font.name or "Calibri", size=cell.font.size or 11,
            bold=True, color="FFFFFF",
        )
        cell.fill = _HEADER_FILL

    percent_like = {"valor)", "Score", "Confidence", "Base", "Bonus", "Penalty"}
    for index, (name, level) in enumerate(plan, start=1):
        column_letter = worksheet.cell(row=1, column=index).column_letter
        if level > 0:
            worksheet.column_dimensions[column_letter].outline_level = level
            worksheet.column_dimensions[column_letter].hidden = True
        max_length = max(
            len(name),
            max(
                (
                    len(str(worksheet.cell(row=r, column=index).value or ""))
                    for r in range(2, min(worksheet.max_row, 200) + 1)
                ),
                default=0,
            ),
        )
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)
        if name.endswith("(percentil)") or name in (
            "Business Score", "Valuation Score", "Financial Score", "Timing Score",
            "Business Confidence", "Valuation Confidence", "Financial Confidence",
            "Timing Confidence", "Investment Score", "Opportunity Score",
            "Conviction Score", "Confidence Score",
        ):
            for r in range(2, worksheet.max_row + 1):
                worksheet.cell(row=r, column=index).number_format = "0.0"

    worksheet.freeze_panes = worksheet.cell(row=2, column=8).coordinate  # após as colunas de identificação
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False


def build_combined_workbook(
    screeners: tuple[tuple[str, str, str], ...],
    output_dir: Path,
    output_path: Path,
    *,
    model_path: Path | None = None,
    deal_breakers_path: Path | None = None,
) -> Path:
    """
    Junta os 3 screeners num único workbook, com a composição completa do
    score (fator por fator, feature por feature) disponível via agrupamento
    de colunas do Excel -- clique no "+" para abrir o detalhe. Reaproveita
    o mesmo pipeline determinístico (scoring.investment.score_dataframe) que
    já produziu os scores finais em research_ranking_report*.json; não
    inventa número novo, só deixa visível o que o pipeline computa e
    descarta no caminho normal. Screener sem coleta feita ainda é
    simplesmente omitido, não é erro.
    """
    model_path = model_path or ROOT / "config" / "model.yaml"
    deal_breakers_path = deal_breakers_path or ROOT / "config" / "deal_breakers.json"
    features_path = model_path.parent / "features.yaml"
    bindings_by_factor = _factor_feature_bindings(features_path, model_path)

    # research_ranking_report*.json/model_portfolio_report*.json são o
    # contrato JSON interno (ver STATUS.md) -- portfolio.model_portfolio os
    # grava em <output_dir>/dados/, não em <output_dir> diretamente.
    data_dir = Path(output_dir) / "dados"

    summary_rows: list[dict[str, Any]] = []
    position_frames: list[pd.DataFrame] = []
    detail_rows: list[dict[str, Any]] = []
    detail_plan: list[tuple[str, int]] | None = None
    scored_cache: dict[str, pd.DataFrame] = {}

    for label, screener_name, collection_relpath in screeners:
        ranking = _load_json(_labeled_path(data_dir, "research_ranking_report.json", label))
        if ranking is None:
            continue
        portfolio = _load_json(_labeled_path(data_dir, "model_portfolio_report.json", label))

        summary_rows.append(_summary_row(screener_name, ranking, portfolio))
        if portfolio is not None:
            position_frames.append(_positions_dataframe(portfolio, screener_name))

        collection_path = ROOT / collection_relpath
        if collection_relpath not in scored_cache:
            scored_cache[collection_relpath] = score_collection(
                collection_path, model_path, deal_breakers_path
            )
        rows, plan = _row_and_groups(
            screener_name, ranking, scored_cache[collection_relpath], bindings_by_factor
        )
        detail_rows.extend(rows)
        detail_plan = plan

    if not summary_rows:
        raise FileNotFoundError(
            "Nenhum research_ranking_report*.json encontrado em "
            f"{data_dir} -- rode portfolio.model_portfolio primeiro."
        )

    summary_df = pd.DataFrame(summary_rows)
    positions_df = (
        pd.concat(position_frames, ignore_index=True) if position_frames else pd.DataFrame()
    )
    column_order = [name for name, _ in (detail_plan or [])]
    detail_df = pd.DataFrame(detail_rows, columns=column_order or None)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Resumo", index=False)
        positions_df.to_excel(writer, sheet_name="Carteira Modelo", index=False)
        detail_df.to_excel(writer, sheet_name="Todos os Itens", index=False)

        _style_flat_sheet(writer, "Resumo", percent_columns=("Peso investido",))
        _style_flat_sheet(writer, "Carteira Modelo", percent_columns=("Peso",))
        _style_grouped_sheet(writer, "Todos os Itens", detail_plan or [])

    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Junta os 3 rankings amplos (S&P 500, Mercado Amplo, ADR) num "
            "único Excel navegável, com a composição do score aberta por "
            "agrupamento de colunas -- não recalcula nada, só expõe o que o "
            "pipeline de scoring já produz."
        )
    )
    parser.add_argument(
        "--output-dir", default=str(ROOT / "output"),
        help=(
            "Diretório raiz de output (contém dados/, onde estão os "
            "research_ranking_report*.json)."
        ),
    )
    parser.add_argument(
        "--output", default=None,
        help="Caminho do arquivo Excel de saída. Default: "
        "<output-dir>/relatorios/research_screeners_combined.xlsx.",
    )
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_path = (
        Path(args.output)
        if args.output
        else output_dir / "relatorios" / "research_screeners_combined.xlsx"
    )
    result = build_combined_workbook(DEFAULT_SCREENERS, output_dir, output_path)
    print(f"Excel combinado gerado em {result}")


if __name__ == "__main__":
    main()
